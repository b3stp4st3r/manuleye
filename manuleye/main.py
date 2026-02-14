import os
import subprocess
import sys
import shutil
import requests
import webbrowser
import time
import tkinter as tk
from tkinter import filedialog
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Prompt
from rich.table import Table
import pyfiglet
import phonenumbers
from phonenumbers import geocoder, carrier, timezone
import re
import json
from datetime import datetime
import socket
import hashlib
import ssl
import zipfile
import tarfile
import random
import string
try:
    import pypdf
except ImportError:
    pypdf = None
try:
    from mutagen.mp3 import MP3
    from mutagen.mp4 import MP4
    from mutagen.easyid3 import EasyID3
except ImportError:
    mutagen = None
try:
    from docx import Document
except ImportError:
    docx_module = None
else:
    docx_module = Document
try:
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None
try:
    import instaloader
except ImportError:
    instaloader = None
console = Console()
REQUIRED_LIBS = ["requests", "rich", "pyfiglet", "Pillow", "phonenumbers", "pypdf", "mutagen", "python-docx", "openpyxl", "instaloader"]
EXTERNAL_TOOLS = ["maigret", "holehe"]
def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')
def get_banner():
    fonts = ["slant", "bloody", "banner3", "doom", "graffiti", "standard"]
    selected_font = random.choice(fonts)
    banner_text = pyfiglet.figlet_format("MANUL - EYE", font=selected_font)
    console.print(f"[bold orange3]{banner_text}[/bold orange3]")
    console.print(f"[dim]Font: {selected_font}[/dim]\n")
def check_requirements():
    clear_screen()
    get_banner()
    console.print(Panel("[bold cyan]Checking Manul's Gear...[/bold cyan]", border_style="cyan"))
    missing_libs = []
    for lib in REQUIRED_LIBS:
        try:
            if lib == "Pillow":
                __import__("PIL")
            elif lib == "pypdf":
                __import__("pypdf")
            elif lib == "python-docx":
                __import__("docx")
            else:
                __import__(lib.lower())
        except ImportError:
            missing_libs.append(lib)
    missing_tools = [tool for tool in EXTERNAL_TOOLS if shutil.which(tool) is None]
    if not missing_libs and not missing_tools:
        console.print("[green][+] All systems working.[/green]")
        time.sleep(1)
        return
    console.print(f"[yellow][!] Missing: {', '.join(missing_libs + missing_tools)}[/yellow]")
    choice = Prompt.ask("\nInstall missing?", choices=["y", "n"], default="y")
    if choice == "y":
        if missing_libs:
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing_libs])
        if "holehe" in missing_tools:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "holehe"])
        if "maigret" in missing_tools:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "maigret"])
        console.print("[bold green][+] Gear updated. Restarting...[/bold green]")
        os.execl(sys.executable, sys.executable, *sys.argv)
def select_file():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_path = filedialog.askopenfilename()
    root.destroy()
    return file_path
def validate_ip(ip):
    ipv4_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
    ipv6_pattern = r'^([0-9a-fA-F]{0,4}:){7}[0-9a-fA-F]{0,4}$'
    if re.match(ipv4_pattern, ip):
        parts = ip.split('.')
        return all(0 <= int(part) <= 255 for part in parts)
    elif re.match(ipv6_pattern, ip):
        return True
    return False
def convert_gps_to_decimal(gps_coords, ref):
    try:
        degrees = float(gps_coords[0])
        minutes = float(gps_coords[1])
        seconds = float(gps_coords[2])
        decimal = degrees + (minutes / 60.0) + (seconds / 3600.0)
        if ref in ['S', 'W']:
            decimal = -decimal
        return decimal
    except:
        return None
def parse_gps_info(gps_data):
    try:
        lat = gps_data.get('GPSLatitude')
        lat_ref = gps_data.get('GPSLatitudeRef')
        lon = gps_data.get('GPSLongitude')
        lon_ref = gps_data.get('GPSLongitudeRef')
        if lat and lon and lat_ref and lon_ref:
            latitude = convert_gps_to_decimal(lat, lat_ref)
            longitude = convert_gps_to_decimal(lon, lon_ref)
            if latitude and longitude:
                return {
                    'latitude': latitude,
                    'longitude': longitude,
                    'google_maps': f"https://www.google.com/maps?q={latitude},{longitude}",
                    'coordinates': f"{latitude}, {longitude}"
                }
    except Exception as e:
        console.print(f"[yellow]GPS parsing error: {e}[/yellow]")
    return None
def local_phone_scan(number):
    console.print(f"[*] Analyzing phone number: {number}...", style="cyan")
    try:
        parsed_num = phonenumbers.parse(number)
        if not phonenumbers.is_valid_number(parsed_num):
            console.print("[red][!] Invalid phone number format![/red]")
            return
        table = Table(title=f"Phone Intelligence: {number}", border_style="orange3")
        table.add_column("Parameter", style="bold white")
        table.add_column("Value", style="green")
        region = geocoder.description_for_number(parsed_num, "ru")
        operator = carrier.name_for_number(parsed_num, "ru")
        zones = timezone.time_zones_for_number(parsed_num)
        table.add_row("Country/Region", region if region else "Unknown")
        table.add_row("Carrier/Provider", operator if operator else "Unknown")
        table.add_row("Time Zones", str(zones))
        table.add_row("International Format", phonenumbers.format_number(parsed_num, phonenumbers.PhoneNumberFormat.INTERNATIONAL))
        table.add_row("Local Format", phonenumbers.format_number(parsed_num, phonenumbers.PhoneNumberFormat.NATIONAL))
        table.add_row("E.164 Format", phonenumbers.format_number(parsed_num, phonenumbers.PhoneNumberFormat.E164))
        console.print(table)
    except Exception as e:
        console.print(f"[red]Error parsing number: {e}[/red]")
def exif_extractor():
    path = select_file()
    if not path: return
    try:
        img = Image.open(path)
        info = img._getexif()
        table = Table(title=f"EXIF: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Tag", style="bold white")
        table.add_column("Value", style="green")
        gps_location = None
        if info:
            for tag, value in info.items():
                tag_name = TAGS.get(tag, tag)
                if tag_name == "GPSInfo":
                    gps_data = {GPSTAGS.get(t, t): value[t] for t in value}
                    gps_location = parse_gps_info(gps_data)
                    if gps_location:
                        table.add_row("GPS Latitude", str(gps_location['latitude']))
                        table.add_row("GPS Longitude", str(gps_location['longitude']))
                        table.add_row("Coordinates", gps_location['coordinates'])
                        table.add_row("Google Maps", gps_location['google_maps'])
                    else:
                        table.add_row("GPS Data", str(gps_data))
                else:
                    table.add_row(str(tag_name), str(value)[:100])
            console.print(table)
            if gps_location:
                open_map = Prompt.ask("\nOpen location in browser?", choices=["y", "n"], default="n")
                if open_map == "y":
                    webbrowser.open(gps_location['google_maps'])
        else:
            console.print("[yellow][!] No metadata found.[/yellow]")
    except FileNotFoundError:
        console.print("[red][!] File not found![/red]")
    except Exception as e:
        console.print(f"[red]Error reading EXIF: {e}[/red]")
    input("\n[Press Enter]")
def get_ip_info(ip):
    if not validate_ip(ip):
        console.print("[red][!] Invalid IP address format![/red]")
        input("\n[Press Enter]")
        return
    try:
        console.print(f"[cyan][*] Querying IP: {ip}...[/cyan]")
        response = requests.get(f"http://ip-api.com/json/{ip}", timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get("status") == "success":
            table = Table(title=f"IP Intel: {ip}", border_style="orange3")
            table.add_column("Parameter", style="bold white")
            table.add_column("Value", style="green")
            key_order = ["country", "regionName", "city", "zip", "lat", "lon", "timezone", "isp", "org", "as"]
            for key in key_order:
                if key in data:
                    table.add_row(key.upper(), str(data[key]))
            console.print(table)
            if "lat" in data and "lon" in data:
                maps_url = f"https://www.google.com/maps?q={data['lat']},{data['lon']}"
                console.print(f"\n[cyan]Google Maps: {maps_url}[/cyan]")
        else:
            console.print(f"[red][!] IP lookup failed: {data.get('message', 'Unknown error')}[/red]")
    except requests.exceptions.Timeout:
        console.print("[red][!] Request timeout. Check your connection.[/red]")
    except requests.exceptions.ConnectionError:
        console.print("[red][!] Network error. Check your internet connection.[/red]")
    except requests.exceptions.RequestException as e:
        console.print(f"[red][!] Request error: {e}[/red]")
    except Exception as e:
        console.print(f"[red][!] Unexpected error: {e}[/red]")
    input("\n[Press Enter]")
def domain_lookup(domain):
    console.print(f"[cyan][*] Analyzing domain: {domain}...[/cyan]")
    try:
        ip = socket.gethostbyname(domain)
        console.print(f"[green][+] Resolved IP: {ip}[/green]\n")
        get_ip_info(ip)
    except socket.gaierror:
        console.print("[red][!] Domain not found or DNS error.[/red]")
        input("\n[Press Enter]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
        input("\n[Press Enter]")
def hash_analyzer():
    path = select_file()
    if not path:
        return
    try:
        console.print(f"[cyan][*] Calculating hashes for: {os.path.basename(path)}...[/cyan]")
        with open(path, 'rb') as f:
            data = f.read()
        table = Table(title=f"File Hashes: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Algorithm", style="bold white")
        table.add_column("Hash", style="green")
        table.add_row("MD5", hashlib.md5(data).hexdigest())
        table.add_row("SHA1", hashlib.sha1(data).hexdigest())
        table.add_row("SHA256", hashlib.sha256(data).hexdigest())
        table.add_row("File Size", f"{len(data)} bytes")
        console.print(table)
        console.print("\n[cyan]Check hash on VirusTotal: https://www.virustotal.com/[/cyan]")
    except FileNotFoundError:
        console.print("[red][!] File not found![/red]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def sqlmap_scan():
    target = Prompt.ask("Target URL (e.g. http://site.com/page.php?id=1)")
    console.print(f"[cyan][*] Starting SQLMap scan on: {target}...[/cyan]")
    console.print("[yellow][!] This may take a while...[/yellow]\n")
    try:
        subprocess.run(f"sqlmap -u {target} --batch --banner", shell=True, timeout=600)
    except subprocess.TimeoutExpired:
        console.print("[red][!] SQLMap timeout (10 min limit)[/red]")
    except FileNotFoundError:
        console.print("[red][!] SQLMap not found. Install: pip install sqlmap-python[/red]")
        console.print("[yellow]Or download from: https://sqlmap.org/[/yellow]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def xss_scan():
    target = Prompt.ask("Target URL")
    console.print(f"[cyan][*] Starting XSStrike scan on: {target}...[/cyan]\n")
    try:
        subprocess.run(f"xsstrike -u {target}", shell=True, timeout=300)
    except subprocess.TimeoutExpired:
        console.print("[red][!] XSStrike timeout (5 min limit)[/red]")
    except FileNotFoundError:
        console.print("[red][!] XSStrike not found. Install: pip install xsstrike[/red]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def metadata_cleaner():
    path = select_file()
    if not path:
        return
    try:
        img = Image.open(path)
        has_exif = bool(img._getexif())
        if has_exif:
            console.print(f"[yellow][!] Found metadata in: {os.path.basename(path)}[/yellow]")
            data = list(img.getdata())
            clean_img = Image.new(img.mode, img.size)
            clean_img.putdata(data)
            base, ext = os.path.splitext(path)
            clean_path = f"{base}_clean{ext}"
            clean_img.save(clean_path)
            console.print(f"[green][+] Cleaned image saved: {clean_path}[/green]")
        else:
            console.print("[green][+] No metadata found. Image is clean.[/green]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def pdf_metadata_extractor():
    path = select_file()
    if not path:
        return
    try:
        if pypdf is None:
            console.print("[red][!] pypdf not installed. Run: pip install pypdf[/red]")
            input("\n[Press Enter]")
            return
        with open(path, 'rb') as f:
            pdf = pypdf.PdfReader(f)
            info = pdf.metadata
            if info:
                table = Table(title=f"PDF Metadata: {os.path.basename(path)}", border_style="orange3")
                table.add_column("Field", style="bold white")
                table.add_column("Value", style="green")
                metadata_fields = {
                    '/Title': 'Title',
                    '/Author': 'Author',
                    '/Subject': 'Subject',
                    '/Creator': 'Creator',
                    '/Producer': 'Producer',
                    '/CreationDate': 'Creation Date',
                    '/ModDate': 'Modification Date'
                }
                for key, label in metadata_fields.items():
                    if key in info:
                        table.add_row(label, str(info[key]))
                table.add_row("Pages", str(len(pdf.pages)))
                console.print(table)
            else:
                console.print("[yellow][!] No metadata found in PDF.[/yellow]")
    except Exception as e:
        console.print(f"[red][!] Error reading PDF: {e}[/red]")
    input("\n[Press Enter]")
def save_results(data, filename=None):
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"manul_results_{timestamp}.json"
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        console.print(f"[green][+] Results saved to: {filename}[/green]")
    except Exception as e:
        console.print(f"[red][!] Save error: {e}[/red]")
def video_metadata_extractor():
    path = select_file()
    if not path:
        return
    try:
        if mutagen is None:
            console.print("[red][!] mutagen not installed. Run: pip install mutagen[/red]")
            input("\n[Press Enter]")
            return
        video = MP4(path)
        table = Table(title=f"Video Metadata: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        if video.tags:
            for key, value in video.tags.items():
                table.add_row(str(key), str(value))
        table.add_row("Duration", f"{video.info.length:.2f} seconds")
        table.add_row("Bitrate", f"{video.info.bitrate} bps")
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading video: {e}[/red]")
    input("\n[Press Enter]")
def audio_metadata_extractor():
    path = select_file()
    if not path:
        return
    try:
        if mutagen is None:
            console.print("[red][!] mutagen not installed. Run: pip install mutagen[/red]")
            input("\n[Press Enter]")
            return
        audio = MP3(path, ID3=EasyID3)
        table = Table(title=f"Audio Metadata: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        tags = ['title', 'artist', 'album', 'date', 'genre', 'composer']
        for tag in tags:
            if tag in audio:
                table.add_row(tag.upper(), str(audio[tag][0]))
        table.add_row("Duration", f"{audio.info.length:.2f} seconds")
        table.add_row("Bitrate", f"{audio.info.bitrate} bps")
        table.add_row("Sample Rate", f"{audio.info.sample_rate} Hz")
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading audio: {e}[/red]")
    input("\n[Press Enter]")
def office_metadata_extractor():
    path = select_file()
    if not path:
        return
    try:
        table = Table(title=f"Office Metadata: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        if path.endswith('.docx'):
            if docx_module is None:
                console.print("[red][!] python-docx not installed. Run: pip install python-docx[/red]")
                input("\n[Press Enter]")
                return
            doc = docx_module(path)
            props = doc.core_properties
            table.add_row("Author", str(props.author))
            table.add_row("Title", str(props.title))
            table.add_row("Subject", str(props.subject))
            table.add_row("Created", str(props.created))
            table.add_row("Modified", str(props.modified))
            table.add_row("Last Modified By", str(props.last_modified_by))
        elif path.endswith('.xlsx'):
            if openpyxl is None:
                console.print("[red][!] openpyxl not installed. Run: pip install openpyxl[/red]")
                input("\n[Press Enter]")
                return
            wb = load_workbook(path)
            props = wb.properties
            table.add_row("Creator", str(props.creator))
            table.add_row("Title", str(props.title))
            table.add_row("Subject", str(props.subject))
            table.add_row("Created", str(props.created))
            table.add_row("Modified", str(props.modified))
            table.add_row("Last Modified By", str(props.lastModifiedBy))
            table.add_row("Sheets", str(len(wb.sheetnames)))
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading document: {e}[/red]")
    input("\n[Press Enter]")
def archive_info():
    path = select_file()
    if not path:
        return
    try:
        table = Table(title=f"Archive Contents: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Filename", style="bold white")
        table.add_column("Size", style="green")
        table.add_column("Compressed", style="yellow")
        if path.endswith('.zip'):
            with zipfile.ZipFile(path, 'r') as zf:
                for info in zf.filelist[:20]:
                    table.add_row(
                        info.filename,
                        f"{info.file_size} bytes",
                        f"{info.compress_size} bytes"
                    )
                if len(zf.filelist) > 20:
                    console.print(f"\n[yellow]Showing 20 of {len(zf.filelist)} files[/yellow]")
        elif path.endswith(('.tar', '.tar.gz', '.tgz')):
            with tarfile.open(path, 'r:*') as tf:
                members = tf.getmembers()[:20]
                for member in members:
                    table.add_row(
                        member.name,
                        f"{member.size} bytes",
                        "N/A"
                    )
                if len(tf.getmembers()) > 20:
                    console.print(f"\n[yellow]Showing 20 of {len(tf.getmembers())} files[/yellow]")
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading archive: {e}[/red]")
    input("\n[Press Enter]")
def port_scanner():
    target = Prompt.ask("Target IP/Domain")
    ports_input = Prompt.ask("Ports (e.g. 80,443 or 1-1000)", default="21,22,23,25,80,443,3306,3389,8080")
    console.print(f"[cyan][*] Scanning {target}...[/cyan]\n")
    try:
        if '-' in ports_input:
            start, end = map(int, ports_input.split('-'))
            ports = range(start, end + 1)
        else:
            ports = [int(p.strip()) for p in ports_input.split(',')]
        open_ports = []
        for port in ports:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex((target, port))
            if result == 0:
                open_ports.append(port)
                console.print(f"[green][+] Port {port} is OPEN[/green]")
            sock.close()
        if not open_ports:
            console.print("[yellow][!] No open ports found[/yellow]")
        else:
            console.print(f"\n[green][+] Found {len(open_ports)} open ports[/green]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def traceroute():
    target = Prompt.ask("Target IP/Domain")
    console.print(f"[cyan][*] Tracing route to {target}...[/cyan]\n")
    try:
        if os.name == 'nt':
            subprocess.run(f"tracert {target}", shell=True)
        else:
            subprocess.run(f"traceroute {target}", shell=True)
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def ping_sweep():
    network = Prompt.ask("Network (e.g. 192.168.1)", default="192.168.1")
    start = int(Prompt.ask("Start IP", default="1"))
    end = int(Prompt.ask("End IP", default="254"))
    console.print(f"[cyan][*] Scanning {network}.{start}-{end}...[/cyan]\n")
    alive_hosts = []
    for i in range(start, end + 1):
        ip = f"{network}.{i}"
        if os.name == 'nt':
            response = os.system(f"ping -n 1 -w 500 {ip} > nul 2>&1")
        else:
            response = os.system(f"ping -c 1 -W 1 {ip} > /dev/null 2>&1")
        if response == 0:
            alive_hosts.append(ip)
            console.print(f"[green][+] {ip} is UP[/green]")
    console.print(f"\n[green][+] Found {len(alive_hosts)} alive hosts[/green]")
    input("\n[Press Enter]")
def ssl_cert_info():
    target = Prompt.ask("Domain (e.g. google.com)")
    port = int(Prompt.ask("Port", default="443"))
    console.print(f"[cyan][*] Fetching SSL certificate from {target}:{port}...[/cyan]\n")
    try:
        context = ssl.create_default_context()
        with socket.create_connection((target, port), timeout=10) as sock:
            with context.wrap_socket(sock, server_hostname=target) as ssock:
                cert = ssock.getpeercert()
        table = Table(title=f"SSL Certificate: {target}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        table.add_row("Subject", str(dict(x[0] for x in cert['subject'])))
        table.add_row("Issuer", str(dict(x[0] for x in cert['issuer'])))
        table.add_row("Version", str(cert['version']))
        table.add_row("Serial Number", str(cert['serialNumber']))
        table.add_row("Not Before", str(cert['notBefore']))
        table.add_row("Not After", str(cert['notAfter']))
        if 'subjectAltName' in cert:
            alt_names = ', '.join([x[1] for x in cert['subjectAltName']])
            table.add_row("Alt Names", alt_names)
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def http_headers_analyzer():
    target = Prompt.ask("URL (e.g. https://example.com)")
    if not target.startswith('http'):
        target = 'https://' + target
    console.print(f"[cyan][*] Fetching headers from {target}...[/cyan]\n")
    try:
        response = requests.get(target, timeout=10, allow_redirects=True)
        table = Table(title=f"HTTP Headers: {target}", border_style="orange3")
        table.add_column("Header", style="bold white")
        table.add_column("Value", style="green")
        table.add_row("Status Code", str(response.status_code))
        for header, value in response.headers.items():
            table.add_row(header, str(value)[:100])
        console.print(table)
        security_headers = ['Strict-Transport-Security', 'Content-Security-Policy', 'X-Frame-Options', 'X-Content-Type-Options']
        missing = [h for h in security_headers if h not in response.headers]
        if missing:
            console.print(f"\n[yellow][!] Missing security headers: {', '.join(missing)}[/yellow]")
        else:
            console.print("\n[green][+] All security headers present[/green]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def instagram_osint():
    username = Prompt.ask("Instagram Username")
    if instaloader is None:
        console.print("[red][!] instaloader not installed. Run: pip install instaloader[/red]")
        input("\n[Press Enter]")
        return
    console.print(f"[cyan][*] Fetching data for @{username}...[/cyan]\n")
    try:
        L = instaloader.Instaloader()
        profile = instaloader.Profile.from_username(L.context, username)
        table = Table(title=f"Instagram: @{username}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        table.add_row("User ID", str(profile.userid))
        table.add_row("Full Name", profile.full_name)
        table.add_row("Biography", profile.biography[:100] if profile.biography else "N/A")
        table.add_row("Followers", str(profile.followers))
        table.add_row("Following", str(profile.followees))
        table.add_row("Posts", str(profile.mediacount))
        table.add_row("Is Private", str(profile.is_private))
        table.add_row("Is Verified", str(profile.is_verified))
        table.add_row("Is Business", str(profile.is_business_account))
        if profile.external_url:
            table.add_row("External URL", profile.external_url)
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    input("\n[Press Enter]")
def password_generator():
    console.print("[bold cyan]═══ PASSWORD GENERATOR ═══[/bold cyan]\n")
    length = int(Prompt.ask("Password length", default="16"))
    count = int(Prompt.ask("Number of passwords", default="5"))
    console.print("\n[yellow]Character types to include:[/yellow]")
    use_uppercase = Prompt.ask("Uppercase letters (A-Z)?", choices=["y", "n"], default="y") == "y"
    use_lowercase = Prompt.ask("Lowercase letters (a-z)?", choices=["y", "n"], default="y") == "y"
    use_digits = Prompt.ask("Digits (0-9)?", choices=["y", "n"], default="y") == "y"
    use_special = Prompt.ask("Special characters (!@#$%^&*)?", choices=["y", "n"], default="y") == "y"
    console.print("\n[yellow]Additional options:[/yellow]")
    exclude_ambiguous = Prompt.ask("Exclude ambiguous chars (0,O,l,1,I)?", choices=["y", "n"], default="n") == "y"
    readable = Prompt.ask("Make more readable (no consecutive special)?", choices=["y", "n"], default="n") == "y"
    charset = ""
    if use_uppercase:
        charset += string.ascii_uppercase
    if use_lowercase:
        charset += string.ascii_lowercase
    if use_digits:
        charset += string.digits
    if use_special:
        charset += "!@#$%^&*()_+-=[]{}|;:,.<>?"
    if exclude_ambiguous:
        ambiguous = "0Ol1I"
        charset = ''.join(c for c in charset if c not in ambiguous)
    if not charset:
        console.print("[red][!] No character types selected![/red]")
        input("\n[Press Enter]")
        return
    passwords = []
    for _ in range(count):
        if readable:
            password = []
            special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
            last_was_special = False
            for i in range(length):
                if last_was_special:
                    char = random.choice([c for c in charset if c not in special_chars])
                    last_was_special = False
                else:
                    char = random.choice(charset)
                    if char in special_chars:
                        last_was_special = True
                password.append(char)
            passwords.append(''.join(password))
        else:
            password = ''.join(random.choice(charset) for _ in range(length))
            passwords.append(password)
    table = Table(title="Generated Passwords", border_style="cyan")
    table.add_column("#", style="dim", width=4)
    table.add_column("Password", style="bold green")
    table.add_column("Strength", style="yellow")
    for i, pwd in enumerate(passwords, 1):
        strength = calculate_password_strength(pwd)
        table.add_row(str(i), pwd, strength)
    console.print("\n")
    console.print(table)
    save = Prompt.ask("\nSave to file?", choices=["y", "n"], default="n")
    if save == "y":
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"passwords_{timestamp}.txt"
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("═" * 60 + "\n")
                f.write("GENERATED PASSWORDS\n")
                f.write(f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Length: {length} | Count: {count}\n")
                f.write("═" * 60 + "\n\n")
                for i, pwd in enumerate(passwords, 1):
                    strength = calculate_password_strength(pwd)
                    f.write(f"{i}. {pwd} [{strength}]\n")
                f.write("\n" + "═" * 60 + "\n")
                f.write("Generated by MANUL-EYE OSINT Framework\n")
                f.write("═" * 60 + "\n")
            console.print(f"[green][+] Passwords saved to: {filename}[/green]")
        except Exception as e:
            console.print(f"[red][!] Error saving: {e}[/red]")
    input("\n[Press Enter]")
def calculate_password_strength(password):
    score = 0
    if len(password) >= 8:
        score += 1
    if len(password) >= 12:
        score += 1
    if len(password) >= 16:
        score += 1
    if any(c.isupper() for c in password):
        score += 1
    if any(c.islower() for c in password):
        score += 1
    if any(c.isdigit() for c in password):
        score += 1
    if any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in password):
        score += 1
    unique_chars = len(set(password))
    if unique_chars >= len(password) * 0.7:
        score += 1
    if score <= 3:
        return "Weak"
    elif score <= 5:
        return "Medium"
    elif score <= 7:
        return "Strong"
    else:
        return "Very Strong"
def create_paste():
    console.print("[bold cyan]═══ PASTE CREATOR ═══[/bold cyan]\n")
    target_name = Prompt.ask("Target Name/Nickname")
    paste_data = {
        "target": target_name,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "personal_info": {},
        "documents": {},
        "social_media": {},
        "contacts": {},
        "locations": {},
        "work_education": {},
        "relatives": [],
        "additional": {},
        "notes": []
    }
    console.print("\n[yellow]Fill in available information (press Enter to skip):[/yellow]\n")
    console.print("[bold orange3]Personal Information:[/bold orange3]")
    paste_data["personal_info"]["full_name"] = Prompt.ask("Full Name (ФИО)", default="")
    paste_data["personal_info"]["date_of_birth"] = Prompt.ask("Date of Birth (ДД.ММ.ГГГГ)", default="")
    paste_data["personal_info"]["age"] = Prompt.ask("Age", default="")
    paste_data["personal_info"]["gender"] = Prompt.ask("Gender", default="")
    console.print("\n[bold orange3]Documents:[/bold orange3]")
    paste_data["documents"]["passport"] = Prompt.ask("Passport (Серия Номер)", default="")
    paste_data["documents"]["inn"] = Prompt.ask("INN (ИНН)", default="")
    paste_data["documents"]["snils"] = Prompt.ask("SNILS (СНИЛС)", default="")
    paste_data["documents"]["ogrnip"] = Prompt.ask("OGRNIP/OGRN (ОГРНИП/ОГРН)", default="")
    console.print("\n[bold orange3]Social Media:[/bold orange3]")
    paste_data["social_media"]["vk"] = Prompt.ask("VK (полная ссылка)", default="")
    paste_data["social_media"]["telegram"] = Prompt.ask("Telegram (@username / ID)", default="")
    paste_data["social_media"]["instagram"] = Prompt.ask("Instagram (полная ссылка)", default="")
    paste_data["social_media"]["tiktok"] = Prompt.ask("TikTok (полная ссылка)", default="")
    paste_data["social_media"]["ok"] = Prompt.ask("Odnoklassniki (полная ссылка)", default="")
    paste_data["social_media"]["twitter"] = Prompt.ask("Twitter/X", default="")
    paste_data["social_media"]["youtube"] = Prompt.ask("YouTube", default="")
    paste_data["social_media"]["other"] = Prompt.ask("Other Social Media", default="")
    console.print("\n[bold orange3]Contacts:[/bold orange3]")
    paste_data["contacts"]["phone_primary"] = Prompt.ask("Phone Primary", default="")
    paste_data["contacts"]["phone_secondary"] = Prompt.ask("Phone Secondary", default="")
    paste_data["contacts"]["email_primary"] = Prompt.ask("Email Primary", default="")
    paste_data["contacts"]["email_secondary"] = Prompt.ask("Email Secondary", default="")
    paste_data["contacts"]["email_other"] = Prompt.ask("Other Emails (comma separated)", default="")
    console.print("\n[bold orange3]Locations:[/bold orange3]")
    paste_data["locations"]["address_primary"] = Prompt.ask("Primary Address (полный адрес)", default="")
    paste_data["locations"]["address_secondary"] = Prompt.ask("Secondary Address", default="")
    paste_data["locations"]["city"] = Prompt.ask("City", default="")
    paste_data["locations"]["region"] = Prompt.ask("Region/Oblast", default="")
    paste_data["locations"]["postal_code"] = Prompt.ask("Postal Code", default="")
    console.print("\n[bold orange3]Work & Education:[/bold orange3]")
    paste_data["work_education"]["workplace"] = Prompt.ask("Workplace (название организации)", default="")
    paste_data["work_education"]["position"] = Prompt.ask("Position (должность)", default="")
    paste_data["work_education"]["work_since"] = Prompt.ask("Working Since (дата)", default="")
    paste_data["work_education"]["work_address"] = Prompt.ask("Work Address (полный адрес)", default="")
    paste_data["work_education"]["work_inn"] = Prompt.ask("Work INN (ИНН организации)", default="")
    paste_data["work_education"]["work_ogrn"] = Prompt.ask("Work OGRN (ОГРН организации)", default="")
    paste_data["work_education"]["school"] = Prompt.ask("School/University", default="")
    paste_data["work_education"]["school_address"] = Prompt.ask("School Address", default="")
    paste_data["work_education"]["education_level"] = Prompt.ask("Education Level", default="")
    console.print("\n[bold orange3]Relatives:[/bold orange3]")
    add_relatives = Prompt.ask("Add relatives info? (y/n)", choices=["y", "n"], default="n")
    if add_relatives == "y":
        while True:
            console.print("\n[cyan]--- Adding Relative ---[/cyan]")
            relative = {}
            relative["relation"] = Prompt.ask("Relation (Мать/Отец/Сестра/Брат/etc)", default="")
            if not relative["relation"]:
                break
            relative["full_name"] = Prompt.ask("Full Name (ФИО)", default="")
            relative["date_of_birth"] = Prompt.ask("Date of Birth", default="")
            relative["phone"] = Prompt.ask("Phone", default="")
            relative["email"] = Prompt.ask("Email", default="")
            relative["vk"] = Prompt.ask("VK", default="")
            relative["telegram"] = Prompt.ask("Telegram", default="")
            relative["instagram"] = Prompt.ask("Instagram", default="")
            relative["address"] = Prompt.ask("Address", default="")
            relative["workplace"] = Prompt.ask("Workplace", default="")
            relative["position"] = Prompt.ask("Position", default="")
            relative["passport"] = Prompt.ask("Passport", default="")
            relative["inn"] = Prompt.ask("INN", default="")
            relative["snils"] = Prompt.ask("SNILS", default="")
            paste_data["relatives"].append(relative)
            more = Prompt.ask("Add another relative? (y/n)", choices=["y", "n"], default="n")
            if more == "n":
                break
    console.print("\n[bold orange3]Additional Info:[/bold orange3]")
    paste_data["additional"]["ip_address"] = Prompt.ask("IP Address", default="")
    paste_data["additional"]["mac_address"] = Prompt.ask("MAC Address", default="")
    paste_data["additional"]["car_number"] = Prompt.ask("Car Number (номер авто)", default="")
    paste_data["additional"]["usernames"] = Prompt.ask("Other Usernames", default="")
    paste_data["additional"]["websites"] = Prompt.ask("Personal Websites", default="")
    paste_data["additional"]["crypto_wallets"] = Prompt.ask("Crypto Wallets", default="")
    console.print("\n[bold orange3]Notes:[/bold orange3]")
    while True:
        note = Prompt.ask("Add note (or press Enter to finish)", default="")
        if not note:
            break
        paste_data["notes"].append(note)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"paste_{target_name.replace(' ', '_')}_{timestamp}.txt"
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("═" * 80 + "\n")
            f.write(f"PASTE: {target_name}\n")
            f.write(f"Created: {paste_data['timestamp']}\n")
            f.write("═" * 80 + "\n\n")
            f.write("【 PERSONAL INFORMATION 】\n")
            f.write("-" * 80 + "\n")
            if paste_data["personal_info"]["full_name"]:
                f.write(f"ФИО: {paste_data['personal_info']['full_name']}\n")
            if paste_data["personal_info"]["date_of_birth"]:
                f.write(f"Дата рождения: {paste_data['personal_info']['date_of_birth']}\n")
            if paste_data["personal_info"]["age"]:
                f.write(f"Возраст: {paste_data['personal_info']['age']}\n")
            if paste_data["personal_info"]["gender"]:
                f.write(f"Пол: {paste_data['personal_info']['gender']}\n")
            f.write("\n")
            if any(paste_data["documents"].values()):
                f.write("【 DOCUMENTS 】\n")
                f.write("-" * 80 + "\n")
                if paste_data["documents"]["passport"]:
                    f.write(f"Паспорт: {paste_data['documents']['passport']}\n")
                if paste_data["documents"]["inn"]:
                    f.write(f"ИНН: {paste_data['documents']['inn']}\n")
                if paste_data["documents"]["snils"]:
                    f.write(f"СНИЛС: {paste_data['documents']['snils']}\n")
                if paste_data["documents"]["ogrnip"]:
                    f.write(f"ОГРН: {paste_data['documents']['ogrnip']}\n")
                f.write("\n")
            f.write("【 CONTACTS 】\n")
            f.write("-" * 80 + "\n")
            if paste_data["contacts"]["phone_primary"]:
                f.write(f"Телефон: {paste_data['contacts']['phone_primary']}\n")
            if paste_data["contacts"]["phone_secondary"]:
                f.write(f"Телефон 2: {paste_data['contacts']['phone_secondary']}\n")
            if paste_data["contacts"]["email_primary"]:
                f.write(f"Эл. почта: {paste_data['contacts']['email_primary']}\n")
            if paste_data["contacts"]["email_secondary"]:
                f.write(f"Эл. почта 2: {paste_data['contacts']['email_secondary']}\n")
            if paste_data["contacts"]["email_other"]:
                f.write(f"Другие почты: {paste_data['contacts']['email_other']}\n")
            f.write("\n")
            f.write("【 SOCIAL MEDIA 】\n")
            f.write("-" * 80 + "\n")
            if paste_data["social_media"]["vk"]:
                f.write(f"ВК: {paste_data['social_media']['vk']}\n")
            if paste_data["social_media"]["telegram"]:
                f.write(f"ТГ: {paste_data['social_media']['telegram']}\n")
            if paste_data["social_media"]["instagram"]:
                f.write(f"Инстаграм: {paste_data['social_media']['instagram']}\n")
            if paste_data["social_media"]["tiktok"]:
                f.write(f"ТТ: {paste_data['social_media']['tiktok']}\n")
            if paste_data["social_media"]["ok"]:
                f.write(f"ОК: {paste_data['social_media']['ok']}\n")
            if paste_data["social_media"]["twitter"]:
                f.write(f"Twitter: {paste_data['social_media']['twitter']}\n")
            if paste_data["social_media"]["youtube"]:
                f.write(f"YouTube: {paste_data['social_media']['youtube']}\n")
            if paste_data["social_media"]["other"]:
                f.write(f"Другое: {paste_data['social_media']['other']}\n")
            f.write("\n")
            f.write("【 LOCATIONS 】\n")
            f.write("-" * 80 + "\n")
            if paste_data["locations"]["address_primary"]:
                f.write(f"Адрес: {paste_data['locations']['address_primary']}\n")
            if paste_data["locations"]["address_secondary"]:
                f.write(f"Адрес 2: {paste_data['locations']['address_secondary']}\n")
            if paste_data["locations"]["city"]:
                f.write(f"Город: {paste_data['locations']['city']}\n")
            if paste_data["locations"]["region"]:
                f.write(f"Регион: {paste_data['locations']['region']}\n")
            if paste_data["locations"]["postal_code"]:
                f.write(f"Индекс: {paste_data['locations']['postal_code']}\n")
            f.write("\n")
            if any(paste_data["work_education"].values()):
                f.write("【 WORK & EDUCATION 】\n")
                f.write("-" * 80 + "\n")
                if paste_data["work_education"]["workplace"]:
                    f.write(f"Место работы: {paste_data['work_education']['workplace']}\n")
                if paste_data["work_education"]["position"]:
                    f.write(f"Должность: {paste_data['work_education']['position']}\n")
                if paste_data["work_education"]["work_since"]:
                    f.write(f"Работает с: {paste_data['work_education']['work_since']}\n")
                if paste_data["work_education"]["work_address"]:
                    f.write(f"Адрес работы: {paste_data['work_education']['work_address']}\n")
                if paste_data["work_education"]["work_inn"]:
                    f.write(f"ИНН организации: {paste_data['work_education']['work_inn']}\n")
                if paste_data["work_education"]["work_ogrn"]:
                    f.write(f"ОГРН организации: {paste_data['work_education']['work_ogrn']}\n")
                if paste_data["work_education"]["school"]:
                    f.write(f"Школа/ВУЗ: {paste_data['work_education']['school']}\n")
                if paste_data["work_education"]["school_address"]:
                    f.write(f"Адрес школы: {paste_data['work_education']['school_address']}\n")
                if paste_data["work_education"]["education_level"]:
                    f.write(f"Образование: {paste_data['work_education']['education_level']}\n")
                f.write("\n")
            if paste_data["relatives"]:
                f.write("【 RELATIVES 】\n")
                f.write("-" * 80 + "\n")
                for rel in paste_data["relatives"]:
                    f.write(f"\n>>> {rel.get('relation', 'Unknown')} <<<\n")
                    if rel.get("full_name"):
                        f.write(f"ФИО: {rel['full_name']}\n")
                    if rel.get("date_of_birth"):
                        f.write(f"Дата рождения: {rel['date_of_birth']}\n")
                    if rel.get("phone"):
                        f.write(f"Телефон: {rel['phone']}\n")
                    if rel.get("email"):
                        f.write(f"Эл. почта: {rel['email']}\n")
                    if rel.get("vk"):
                        f.write(f"ВК: {rel['vk']}\n")
                    if rel.get("telegram"):
                        f.write(f"ТГ: {rel['telegram']}\n")
                    if rel.get("instagram"):
                        f.write(f"Инстаграм: {rel['instagram']}\n")
                    if rel.get("address"):
                        f.write(f"Адрес: {rel['address']}\n")
                    if rel.get("workplace"):
                        f.write(f"Место работы: {rel['workplace']}\n")
                    if rel.get("position"):
                        f.write(f"Должность: {rel['position']}\n")
                    if rel.get("passport"):
                        f.write(f"Паспорт: {rel['passport']}\n")
                    if rel.get("inn"):
                        f.write(f"ИНН: {rel['inn']}\n")
                    if rel.get("snils"):
                        f.write(f"СНИЛС: {rel['snils']}\n")
                f.write("\n")
            if any(paste_data["additional"].values()):
                f.write("【 ADDITIONAL INFO 】\n")
                f.write("-" * 80 + "\n")
                for key, value in paste_data["additional"].items():
                    if value:
                        f.write(f"{key.replace('_', ' ').title()}: {value}\n")
                f.write("\n")
            if paste_data["notes"]:
                f.write("【 NOTES 】\n")
                f.write("-" * 80 + "\n")
                for i, note in enumerate(paste_data["notes"], 1):
                    f.write(f"{i}. {note}\n")
                f.write("\n")
            f.write("═" * 80 + "\n")
            f.write("Generated by MANUL-EYE OSINT Framework\n")
            f.write("═" * 80 + "\n")
        json_filename = filename.replace('.txt', '.json')
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(paste_data, f, indent=4, ensure_ascii=False)
        console.print(f"\n[green][+] Paste created successfully![/green]")
        console.print(f"[cyan]Text version: {filename}[/cyan]")
        console.print(f"[cyan]JSON version: {json_filename}[/cyan]")
    except Exception as e:
        console.print(f"[red][!] Error creating paste: {e}[/red]")
    input("\n[Press Enter]")
def show_category_menu():
    while True:
        clear_screen()
        get_banner()
        menu_table = Table(show_header=False, box=None, padding=(0, 2))
        menu_table.add_column(style="white", width=22)
        menu_table.add_column(style="white", width=22)
        menu_table.add_column(style="white", width=22)
        menu_table.add_column(style="white", width=22)
        menu_table.add_row(
            "[bold orange3][LOOKUP][/bold orange3]",
            "[bold red][EXPLOIT][/bold red]",
            "[bold yellow][METADATA][/bold yellow]",
            "[bold cyan][NETWORK][/bold cyan]"
        )
        menu_table.add_row(
            "[1] Username Hunt",
            "[6] SQLMap",
            "[11] EXIF Extract",
            "[16] Port Scanner"
        )
        menu_table.add_row(
            "[2] Email Lookup",
            "[7] XSStrike",
            "[12] PDF Metadata",
            "[17] Traceroute"
        )
        menu_table.add_row(
            "[3] Phone Intel",
            "",
            "[13] Video Metadata",
            "[18] Ping Sweep"
        )
        menu_table.add_row(
            "[4] IP Geolocate",
            "",
            "[14] Audio Metadata",
            "[19] SSL Cert Info"
        )
        menu_table.add_row(
            "[5] Domain Resolve",
            "",
            "[15] Office Metadata",
            "[20] HTTP Headers"
        )
        menu_table.add_row(
            "[8] Instagram OSINT",
            "",
            "[9] Archive Info",
            "[21] Create Paste"
        )
        menu_table.add_row(
            "",
            "",
            "[10] Meta Cleaner",
            "[22] Password Gen"
        )
        console.print(menu_table)
        console.print("\n[bold red][0] Exit[/bold red]")
        choice = Prompt.ask("\n[bold orange3]Select option[/bold orange3]")
        if choice == "1":
            target = Prompt.ask("Username")
            try:
                result = subprocess.run(f"maigret {target} -a", shell=True, capture_output=True, text=True, timeout=300)
                if result.returncode != 0:
                    console.print("[yellow][!] Maigret finished with warnings[/yellow]")
            except subprocess.TimeoutExpired:
                console.print("[red][!] Timeout (5 min limit)[/red]")
            except FileNotFoundError:
                console.print("[red][!] Maigret not found. Install: pip install maigret[/red]")
            except Exception as e:
                console.print(f"[red][!] Error: {e}[/red]")
            input("\n[Press Enter]")
        elif choice == "2":
            target = Prompt.ask("Email")
            try:
                subprocess.run(f"holehe {target} --only-used", shell=True, timeout=60)
            except subprocess.TimeoutExpired:
                console.print("[red][!] Timeout[/red]")
            except FileNotFoundError:
                console.print("[red][!] Holehe not found. Install: pip install holehe[/red]")
            except Exception as e:
                console.print(f"[red][!] Error: {e}[/red]")
            input("\n[Press Enter]")
        elif choice == "3":
            target = Prompt.ask("Phone (e.g. +79991234567)")
            local_phone_scan(target)
            input("\n[Press Enter]")
        elif choice == "4":
            target = Prompt.ask("IP Address")
            get_ip_info(target)
        elif choice == "5":
            target = Prompt.ask("Domain (e.g. example.com)")
            domain_lookup(target)
        elif choice == "6":
            sqlmap_scan()
        elif choice == "7":
            xss_scan()
        elif choice == "8":
            instagram_osint()
        elif choice == "9":
            archive_info()
        elif choice == "10":
            metadata_cleaner()
        elif choice == "11":
            exif_extractor()
        elif choice == "12":
            pdf_metadata_extractor()
        elif choice == "13":
            video_metadata_extractor()
        elif choice == "14":
            audio_metadata_extractor()
        elif choice == "15":
            office_metadata_extractor()
        elif choice == "16":
            port_scanner()
        elif choice == "17":
            traceroute()
        elif choice == "18":
            ping_sweep()
        elif choice == "19":
            ssl_cert_info()
        elif choice == "20":
            http_headers_analyzer()
        elif choice == "21":
            create_paste()
        elif choice == "22":
            password_generator()
        elif choice == "0":
            console.print("[bold cyan]Stay sharp. Stay hidden.[/bold cyan]")
            break
        else:
            console.print("[red][!] Invalid option[/red]")
            time.sleep(1)
def main():
    """
    Main entry point for MANUL-EYE OSINT Framework.
    
    This function serves as the command-line interface entry point.
    It checks system requirements and launches the interactive menu.
    """
    try:
        check_requirements()
        show_category_menu()
    except KeyboardInterrupt:
        console.print("\n[yellow][!] Operation cancelled by user.[/yellow]")
        sys.exit(0)
    except Exception as e:
        console.print(f"\n[red][!] Unexpected error: {e}[/red]")
        sys.exit(1)


if __name__ == "__main__":
    main()