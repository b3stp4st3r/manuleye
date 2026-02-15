import os
import webbrowser
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS
from rich.table import Table
from rich.prompt import Prompt
from manuleye.core import console, select_file

try:
    import pypdf
except ImportError:
    pypdf = None

try:
    from mutagen.mp3 import MP3
    from mutagen.mp4 import MP4
    from mutagen.easyid3 import EasyID3
except ImportError:
    mutagen = None

try:
    from docx import Document
except ImportError:
    docx_module = None
else:
    docx_module = Document

try:
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None

import zipfile
import tarfile


def convert_gps_to_decimal(gps_coords, ref):
    try:
        degrees = float(gps_coords[0])
        minutes = float(gps_coords[1])
        seconds = float(gps_coords[2])
        decimal = degrees + (minutes / 60.0) + (seconds / 3600.0)
        if ref in ['S', 'W']:
            decimal = -decimal
        return decimal
    except:
        return None

def parse_gps_info(gps_data):
    try:
        lat = gps_data.get('GPSLatitude')
        lat_ref = gps_data.get('GPSLatitudeRef')
        lon = gps_data.get('GPSLongitude')
        lon_ref = gps_data.get('GPSLongitudeRef')
        if lat and lon and lat_ref and lon_ref:
            latitude = convert_gps_to_decimal(lat, lat_ref)
            longitude = convert_gps_to_decimal(lon, lon_ref)
            if latitude and longitude:
                return {
                    'latitude': latitude,
                    'longitude': longitude,
                    'google_maps': f"https://www.google.com/maps?q={latitude},{longitude}",
                    'coordinates': f"{latitude}, {longitude}"
                }
    except Exception as e:
        console.print(f"[yellow]GPS parsing error: {e}[/yellow]")
    return None


def exif_extractor():
    path = select_file()
    if not path:
        return
    try:
        img = Image.open(path)
        info = img._getexif()
        table = Table(title=f"EXIF: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Tag", style="bold white")
        table.add_column("Value", style="green")
        
        gps_location = None
        if info:
            for tag, value in info.items():
                tag_name = TAGS.get(tag, tag)
                if tag_name == "GPSInfo":
                    gps_data = {GPSTAGS.get(t, t): value[t] for t in value}
                    gps_location = parse_gps_info(gps_data)
                    if gps_location:
                        table.add_row("GPS Latitude", str(gps_location['latitude']))
                        table.add_row("GPS Longitude", str(gps_location['longitude']))
                        table.add_row("Coordinates", gps_location['coordinates'])
                        table.add_row("Google Maps", gps_location['google_maps'])
                    else:
                        table.add_row("GPS Data", str(gps_data))
                else:
                    table.add_row(str(tag_name), str(value)[:100])
            console.print(table)
            
            if gps_location:
                open_map = Prompt.ask("\nOpen location in browser?", choices=["y", "n"], default="n")
                if open_map == "y":
                    webbrowser.open(gps_location['google_maps'])
        else:
            console.print("[yellow][!] No metadata found.[/yellow]")
    except FileNotFoundError:
        console.print("[red][!] File not found![/red]")
    except Exception as e:
        console.print(f"[red]Error reading EXIF: {e}[/red]")
    
    input("\n[Press Enter]")


def pdf_metadata():
    path = select_file()
    if not path:
        return
    try:
        if pypdf is None:
            console.print("[red][!] pypdf not installed. Run: pip install pypdf[/red]")
            input("\n[Press Enter]")
            return
        
        with open(path, 'rb') as f:
            pdf = pypdf.PdfReader(f)
            info = pdf.metadata
            
            if info:
                table = Table(title=f"PDF Metadata: {os.path.basename(path)}", border_style="orange3")
                table.add_column("Field", style="bold white")
                table.add_column("Value", style="green")
                
                metadata_fields = {
                    '/Title': 'Title',
                    '/Author': 'Author',
                    '/Subject': 'Subject',
                    '/Creator': 'Creator',
                    '/Producer': 'Producer',
                    '/CreationDate': 'Creation Date',
                    '/ModDate': 'Modification Date'
                }
                
                for key, label in metadata_fields.items():
                    if key in info:
                        table.add_row(label, str(info[key]))
                
                table.add_row("Pages", str(len(pdf.pages)))
                console.print(table)
            else:
                console.print("[yellow][!] No metadata found in PDF.[/yellow]")
    except Exception as e:
        console.print(f"[red][!] Error reading PDF: {e}[/red]")
    
    input("\n[Press Enter]")


def video_metadata():
    path = select_file()
    if not path:
        return
    try:
        if mutagen is None:
            console.print("[red][!] mutagen not installed. Run: pip install mutagen[/red]")
            input("\n[Press Enter]")
            return
        
        video = MP4(path)
        table = Table(title=f"Video Metadata: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        
        if video.tags:
            for key, value in video.tags.items():
                table.add_row(str(key), str(value))
        
        table.add_row("Duration", f"{video.info.length:.2f} seconds")
        table.add_row("Bitrate", f"{video.info.bitrate} bps")
        
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading video: {e}[/red]")
    
    input("\n[Press Enter]")


def audio_metadata():
    path = select_file()
    if not path:
        return
    try:
        if mutagen is None:
            console.print("[red][!] mutagen not installed. Run: pip install mutagen[/red]")
            input("\n[Press Enter]")
            return
        
        audio = MP3(path, ID3=EasyID3)
        table = Table(title=f"Audio Metadata: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        
        tags = ['title', 'artist', 'album', 'date', 'genre', 'composer']
        for tag in tags:
            if tag in audio:
                table.add_row(tag.upper(), str(audio[tag][0]))
        
        table.add_row("Duration", f"{audio.info.length:.2f} seconds")
        table.add_row("Bitrate", f"{audio.info.bitrate} bps")
        table.add_row("Sample Rate", f"{audio.info.sample_rate} Hz")
        
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading audio: {e}[/red]")
    
    input("\n[Press Enter]")


def office_metadata():
    path = select_file()
    if not path:
        return
    try:
        table = Table(title=f"Office Metadata: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Field", style="bold white")
        table.add_column("Value", style="green")
        
        if path.endswith('.docx'):
            if docx_module is None:
                console.print("[red][!] python-docx not installed. Run: pip install python-docx[/red]")
                input("\n[Press Enter]")
                return
            
            doc = docx_module(path)
            props = doc.core_properties
            
            table.add_row("Author", str(props.author))
            table.add_row("Title", str(props.title))
            table.add_row("Subject", str(props.subject))
            table.add_row("Created", str(props.created))
            table.add_row("Modified", str(props.modified))
            table.add_row("Last Modified By", str(props.last_modified_by))
            
        elif path.endswith('.xlsx'):
            if openpyxl is None:
                console.print("[red][!] openpyxl not installed. Run: pip install openpyxl[/red]")
                input("\n[Press Enter]")
                return
            
            wb = load_workbook(path)
            props = wb.properties
            
            table.add_row("Creator", str(props.creator))
            table.add_row("Title", str(props.title))
            table.add_row("Subject", str(props.subject))
            table.add_row("Created", str(props.created))
            table.add_row("Modified", str(props.modified))
            table.add_row("Last Modified By", str(props.lastModifiedBy))
            table.add_row("Sheets", str(len(wb.sheetnames)))
        
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading document: {e}[/red]")
    
    input("\n[Press Enter]")


def archive_info():
    path = select_file()
    if not path:
        return
    try:
        table = Table(title=f"Archive Contents: {os.path.basename(path)}", border_style="orange3")
        table.add_column("Filename", style="bold white")
        table.add_column("Size", style="green")
        table.add_column("Compressed", style="yellow")
        
        if path.endswith('.zip'):
            with zipfile.ZipFile(path, 'r') as zf:
                for info in zf.filelist[:20]:
                    table.add_row(
                        info.filename,
                        f"{info.file_size} bytes",
                        f"{info.compress_size} bytes"
                    )
                if len(zf.filelist) > 20:
                    console.print(f"\n[yellow]Showing 20 of {len(zf.filelist)} files[/yellow]")
                    
        elif path.endswith(('.tar', '.tar.gz', '.tgz')):
            with tarfile.open(path, 'r:*') as tf:
                members = tf.getmembers()[:20]
                for member in members:
                    table.add_row(
                        member.name,
                        f"{member.size} bytes",
                        "N/A"
                    )
                if len(tf.getmembers()) > 20:
                    console.print(f"\n[yellow]Showing 20 of {len(tf.getmembers())} files[/yellow]")
        
        console.print(table)
    except Exception as e:
        console.print(f"[red][!] Error reading archive: {e}[/red]")
    
    input("\n[Press Enter]")


def metadata_cleaner():
    path = select_file()
    if not path:
        return
    try:
        img = Image.open(path)
        has_exif = bool(img._getexif())
        
        if has_exif:
            console.print(f"[yellow][!] Found metadata in: {os.path.basename(path)}[/yellow]")
            data = list(img.getdata())
            clean_img = Image.new(img.mode, img.size)
            clean_img.putdata(data)
            
            base, ext = os.path.splitext(path)
            clean_path = f"{base}_clean{ext}"
            clean_img.save(clean_path)
            
            console.print(f"[green][+] Cleaned image saved: {clean_path}[/green]")
        else:
            console.print("[green][+] No metadata found. Image is clean.[/green]")
    except Exception as e:
        console.print(f"[red][!] Error: {e}[/red]")
    
    input("\n[Press Enter]")
